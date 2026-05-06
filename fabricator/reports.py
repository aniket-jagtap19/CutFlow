"""
CutFlow – Report Generators
Produces PDF quotations and Excel BOQ / bar-optimisation reports.
"""

import io
import math
from typing import List, Dict, Optional
from datetime import date

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .calculator import (
    WindowResult, BarUsage, PROFILES, GLASS_OPTIONS,
    FINISH_OPTIONS, TYPOLOGY_LABELS, aggregate_hardware
)


# ─── Colour Palette ───────────────────────────────────────────────────────────
DARK  = colors.HexColor("#0D1B2A")
GOLD  = colors.HexColor("#C9922A")
MID   = colors.HexColor("#1E3A5F")
LIGHT = colors.HexColor("#EDF2F7")
WHITE = colors.white
GREY  = colors.HexColor("#4A5568")

# ─── PDF Report ───────────────────────────────────────────────────────────────

def _header_style():
    s = getSampleStyleSheet()
    return ParagraphStyle("h", parent=s["Normal"],
                          fontSize=22, textColor=WHITE,
                          fontName="Helvetica-Bold", alignment=TA_LEFT)

def _sub_style():
    s = getSampleStyleSheet()
    return ParagraphStyle("sub", parent=s["Normal"],
                          fontSize=10, textColor=GOLD,
                          fontName="Helvetica-Bold", alignment=TA_LEFT)

def _body_style():
    s = getSampleStyleSheet()
    return ParagraphStyle("body", parent=s["Normal"],
                          fontSize=9, textColor=DARK,
                          fontName="Helvetica", leading=13)

def _table_style(header_bg=MID):
    return TableStyle([
        ("BACKGROUND",  (0, 0), (-1, 0), header_bg),
        ("TEXTCOLOR",   (0, 0), (-1, 0), WHITE),
        ("FONTNAME",    (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",    (0, 0), (-1, 0), 9),
        ("ALIGN",       (0, 0), (-1, 0), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [WHITE, LIGHT]),
        ("FONTNAME",    (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE",    (0, 1), (-1, -1), 8),
        ("ALIGN",       (1, 1), (-1, -1), "CENTER"),
        ("GRID",        (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E0")),
        ("TOPPADDING",  (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING",(0,0), (-1, -1), 5),
    ])


def generate_pdf_quotation(results: List[WindowResult],
                           bar_data: Dict[str, List[BarUsage]],
                           company_name: str = "D Sign Design") -> bytes:
    """
    Generate a full PDF quotation covering:
    1. Cover / summary
    2. Per-window profile cut list
    3. Glass schedule
    4. Hardware BOQ
    5. Bar optimisation summary
    6. Grand total quotation
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=15*mm, bottomMargin=15*mm,
                            leftMargin=18*mm, rightMargin=18*mm)
    elems = []
    body  = _body_style()
    sub   = _sub_style()

    # ── Cover Header ──
    cover_data = [[Paragraph(
        f"<b>CutFlow</b> &nbsp;|&nbsp; Window & Door Quotation<br/>"
        f"<font size='10' color='#C9922A'>{company_name} &nbsp;·&nbsp; {date.today().strftime('%d %B %Y')}</font>",
        ParagraphStyle("cov", fontSize=16, textColor=WHITE,
                       fontName="Helvetica-Bold", leading=22))]]
    cov_tbl = Table(cover_data, colWidths=[174*mm])
    cov_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), DARK),
        ("TOPPADDING", (0, 0), (-1, -1), 12),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 12),
        ("LEFTPADDING", (0, 0), (-1, -1), 10),
    ]))
    elems.append(cov_tbl)
    elems.append(Spacer(1, 8*mm))

    # ── 1. Summary Table ──
    elems.append(Paragraph("1. ORDER SUMMARY", sub))
    elems.append(Spacer(1, 2*mm))
    total_units = sum(r.entry.qty for r in results)
    grand_total = sum(r.total_cost for r in results)
    sum_data = [["Code", "Typology", "W×H (mm)", "Qty", "Finish", "Glass", "Mesh", "Unit Cost (₹)", "Total (₹)"]]
    for r in results:
        e = r.entry
        sum_data.append([
            e.code,
            TYPOLOGY_LABELS.get(e.typology, e.typology),
            f"{int(e.width)} × {int(e.height)}",
            str(e.qty),
            FINISH_OPTIONS.get(e.finish, {}).get("label", e.finish),
            GLASS_OPTIONS.get(e.glass_type, {}).get("label", e.glass_type),
            "Yes" if e.mesh else "No",
            f"₹ {r.total_cost/e.qty:,.0f}",
            f"₹ {r.total_cost:,.0f}",
        ])
    sum_data.append(["", "", "", str(total_units), "", "", "", "GRAND TOTAL",
                     f"₹ {grand_total:,.0f}"])
    t = Table(sum_data, colWidths=[w*mm for w in [18, 35, 25, 10, 25, 28, 12, 28, 28]])
    ts = _table_style()
    ts.add("BACKGROUND", (0, len(sum_data)-1), (-1, -1), GOLD)
    ts.add("TEXTCOLOR",  (0, len(sum_data)-1), (-1, -1), WHITE)
    ts.add("FONTNAME",   (0, len(sum_data)-1), (-1, -1), "Helvetica-Bold")
    t.setStyle(ts)
    elems.append(t)
    elems.append(Spacer(1, 6*mm))

    # ── 2. Profile Cut List ──
    elems.append(Paragraph("2. PROFILE CUT LIST", sub))
    elems.append(Spacer(1, 2*mm))
    cut_data = [["Code", "Profile", "Cut Length (mm)", "Pcs/Unit", "Total Pcs", "Total Length (m)"]]
    for r in results:
        for pc in r.profile_cuts:
            total_pcs = pc.count * r.entry.qty
            total_m   = round(pc.length * total_pcs / 1000, 3)
            cut_data.append([r.entry.code, pc.label, f"{pc.length:.0f}",
                             str(pc.count), str(total_pcs), f"{total_m:.3f}"])
    t2 = Table(cut_data, colWidths=[w*mm for w in [22, 45, 30, 22, 22, 33]])
    t2.setStyle(_table_style())
    elems.append(t2)
    elems.append(Spacer(1, 6*mm))

    # ── 3. Glass Schedule ──
    elems.append(Paragraph("3. GLASS SCHEDULE", sub))
    elems.append(Spacer(1, 2*mm))
    glass_data = [["Code", "Glass Type", "W (mm)", "H (mm)", "Area/Unit (m²)", "Qty", "Total Area (m²)"]]
    for r in results:
        e = r.entry
        gl = GLASS_OPTIONS.get(e.glass_type, {}).get("label", e.glass_type)
        glass_data.append([
            e.code, gl,
            f"{r.glass_width:.0f}", f"{r.glass_height:.0f}",
            f"{r.glass_area:.4f}", str(e.qty),
            f"{r.glass_area * e.qty:.4f}",
        ])
    t3 = Table(glass_data, colWidths=[w*mm for w in [20, 38, 22, 22, 30, 15, 30]])
    t3.setStyle(_table_style())
    elems.append(t3)
    elems.append(Spacer(1, 6*mm))

    # ── 4. Hardware BOQ ──
    elems.append(Paragraph("4. HARDWARE & ACCESSORIES BOQ", sub))
    elems.append(Spacer(1, 2*mm))
    agg_hw = aggregate_hardware(results)
    hw_data = [["Item", "Total Qty", "Unit Cost (₹)", "Total Cost (₹)"]]
    from .calculator import HARDWARE_COST
    hw_total = 0
    for item, qty in sorted(agg_hw.items()):
        uc = HARDWARE_COST.get(item, 200)
        tc = uc * qty
        hw_total += tc
        hw_data.append([item, str(qty), f"₹ {uc:,}", f"₹ {tc:,}"])
    hw_data.append(["", "", "TOTAL", f"₹ {hw_total:,}"])
    t4 = Table(hw_data, colWidths=[w*mm for w in [80, 30, 35, 35]])
    ts4 = _table_style()
    ts4.add("BACKGROUND", (0, len(hw_data)-1), (-1, -1), GOLD)
    ts4.add("TEXTCOLOR",  (0, len(hw_data)-1), (-1, -1), WHITE)
    ts4.add("FONTNAME",   (0, len(hw_data)-1), (-1, -1), "Helvetica-Bold")
    t4.setStyle(ts4)
    elems.append(t4)
    elems.append(Spacer(1, 6*mm))

    # ── 5. Bar Optimisation Summary ──
    elems.append(Paragraph("5. BAR OPTIMISATION SUMMARY", sub))
    elems.append(Spacer(1, 2*mm))
    bar_sum = [["Profile", "Bars Req.", "Total Bar Len (m)", "Material Used (m)", "Waste (m)", "Utilisation %"]]
    for pk, bars in bar_data.items():
        if not bars:
            continue
        n_bars    = len(bars)
        total_bar = sum(b.bar_length for b in bars) / 1000
        total_used= sum(b.used for b in bars) / 1000
        total_waste=sum(b.waste for b in bars) / 1000
        avg_util  = (total_used / total_bar * 100) if total_bar else 0
        bar_sum.append([PROFILES[pk]["label"],
                        str(n_bars),
                        f"{total_bar:.2f}",
                        f"{total_used:.2f}",
                        f"{total_waste:.2f}",
                        f"{avg_util:.1f}%"])
    t5 = Table(bar_sum, colWidths=[w*mm for w in [45, 22, 35, 35, 25, 30]])
    t5.setStyle(_table_style())
    elems.append(t5)
    elems.append(Spacer(1, 6*mm))

    # ── 6. Cost Breakdown ──
    elems.append(Paragraph("6. COST BREAKDOWN", sub))
    elems.append(Spacer(1, 2*mm))
    total_profile  = sum(r.profile_cost for r in results)
    total_glass    = sum(r.glass_cost   for r in results)
    total_hw       = sum(r.hardware_cost for r in results)
    total_finish   = sum(r.finish_surcharge for r in results)
    total_all      = sum(r.total_cost   for r in results)
    cost_data = [
        ["Cost Head", "Amount (₹)"],
        ["Aluminium Profiles", f"₹ {total_profile:,.0f}"],
        ["Glass Supply", f"₹ {total_glass:,.0f}"],
        ["Hardware & Accessories", f"₹ {total_hw:,.0f}"],
        ["Finish Surcharge", f"₹ {total_finish:,.0f}"],
        ["GST @ 18%", f"₹ {total_all * 0.18:,.0f}"],
        ["GRAND TOTAL (incl. GST)", f"₹ {total_all * 1.18:,.0f}"],
    ]
    t6 = Table(cost_data, colWidths=[w*mm for w in [100, 74]])
    ts6 = _table_style(DARK)
    ts6.add("BACKGROUND", (0, len(cost_data)-1), (-1, -1), GOLD)
    ts6.add("TEXTCOLOR",  (0, len(cost_data)-1), (-1, -1), WHITE)
    ts6.add("FONTNAME",   (0, len(cost_data)-1), (-1, -1), "Helvetica-Bold")
    t6.setStyle(ts6)
    elems.append(t6)
    elems.append(Spacer(1, 4*mm))
    elems.append(Paragraph(
        "Terms: Prices valid 30 days. 50% advance, balance before delivery. Delivery 4–6 weeks. "
        "Installation not included unless separately quoted.", body))

    doc.build(elems)
    return buf.getvalue()


# ─── Excel Report ─────────────────────────────────────────────────────────────

def _xl_header(ws, row, cols, text, bg="0D1B2A", fg="FFFFFF", bold=True):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=bold, color=fg, size=12 if bold else 10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22

def _xl_row_header(ws, row, headers, bg="1E3A5F"):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 18

def _xl_data_row(ws, row, values, shade=False):
    bg = "EDF2F7" if shade else "FFFFFF"
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.fill = PatternFill("solid", fgColor=bg)
        c.font = Font(size=9)
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 16

def _xl_total_row(ws, row, values, cols, bg="C9922A"):
    for col, v in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=v)
        c.font = Font(bold=True, color="FFFFFF", size=9)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")


def generate_excel_report(results: List[WindowResult],
                           bar_data: Dict[str, List[BarUsage]]) -> bytes:
    """
    Generate multi-sheet Excel workbook:
    Sheet 1 – Summary
    Sheet 2 – Profile Cut List
    Sheet 3 – Glass Schedule
    Sheet 4 – Hardware BOQ
    Sheet 5 – Bar Optimisation (detail)
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Summary")
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 8
    ws1.column_dimensions["E"].width = 16
    ws1.column_dimensions["F"].width = 18
    ws1.column_dimensions["G"].width = 8
    ws1.column_dimensions["H"].width = 16
    ws1.column_dimensions["I"].width = 16

    _xl_header(ws1, 1, 9, "CutFlow – Window & Door Quotation Summary")
    _xl_row_header(ws1, 2, ["Code","Typology","W×H (mm)","Qty","Finish","Glass","Mesh","Unit Cost","Total Cost"])
    for i, r in enumerate(results):
        e = r.entry
        shade = (i % 2 == 1)
        _xl_data_row(ws1, 3+i, [
            e.code,
            TYPOLOGY_LABELS.get(e.typology, e.typology),
            f"{int(e.width)} × {int(e.height)}",
            e.qty,
            FINISH_OPTIONS.get(e.finish, {}).get("label", e.finish),
            GLASS_OPTIONS.get(e.glass_type, {}).get("label", e.glass_type),
            "Yes" if e.mesh else "No",
            r.total_cost / e.qty,
            r.total_cost,
        ], shade)
    grand = sum(r.total_cost for r in results)
    _xl_total_row(ws1, 3+len(results), ["","","",
                  sum(r.entry.qty for r in results),"","","","GRAND TOTAL", grand], 9)
    # Format currency cols
    for row in ws1.iter_rows(min_row=3, max_row=3+len(results), min_col=8, max_col=9):
        for cell in row:
            cell.number_format = '₹ #,##0.00'

    # ── Sheet 2: Profile Cut List ──────────────────────────────────────────────
    ws2 = wb.create_sheet("Profile Cut List")
    for col, w in zip("ABCDEF", [12,30,20,14,14,18]):
        ws2.column_dimensions[col].width = w
    _xl_header(ws2, 1, 6, "PROFILE CUT LIST – All Windows")
    _xl_row_header(ws2, 2, ["Code","Profile","Cut Length (mm)","Pcs/Unit","Total Pcs","Total Length (m)"])
    row = 3
    for r in results:
        for pc in r.profile_cuts:
            total_pcs = pc.count * r.entry.qty
            total_m   = round(pc.length * total_pcs / 1000, 3)
            shade = (row % 2 == 0)
            _xl_data_row(ws2, row, [r.entry.code, pc.label, pc.length,
                                    pc.count, total_pcs, total_m], shade)
            row += 1

    # ── Sheet 3: Glass Schedule ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Glass Schedule")
    for col, w in zip("ABCDEFG", [12,28,14,14,18,8,18]):
        ws3.column_dimensions[col].width = w
    _xl_header(ws3, 1, 7, "GLASS SCHEDULE")
    _xl_row_header(ws3, 2, ["Code","Glass Type","Width (mm)","Height (mm)","Area/Unit (m²)","Qty","Total Area (m²)"])
    for i, r in enumerate(results):
        e = r.entry
        gl = GLASS_OPTIONS.get(e.glass_type, {}).get("label", e.glass_type)
        shade = (i % 2 == 1)
        _xl_data_row(ws3, 3+i, [
            e.code, gl, round(r.glass_width), round(r.glass_height),
            round(r.glass_area, 4), e.qty, round(r.glass_area * e.qty, 4)
        ], shade)

    # ── Sheet 4: Hardware BOQ ─────────────────────────────────────────────────
    ws4 = wb.create_sheet("Hardware BOQ")
    for col, w in zip("ABCD", [35,12,16,16]):
        ws4.column_dimensions[col].width = w
    _xl_header(ws4, 1, 4, "HARDWARE & ACCESSORIES BOQ")
    _xl_row_header(ws4, 2, ["Item","Qty","Unit Cost (₹)","Total Cost (₹)"])
    from .calculator import HARDWARE_COST
    agg = aggregate_hardware(results)
    hw_tot = 0
    for i, (item, qty) in enumerate(sorted(agg.items())):
        uc = HARDWARE_COST.get(item, 200)
        tc = uc * qty
        hw_tot += tc
        shade = (i % 2 == 1)
        _xl_data_row(ws4, 3+i, [item, qty, uc, tc], shade)
    _xl_total_row(ws4, 3+len(agg), ["TOTAL","",""  , hw_tot], 4)

    # ── Sheet 5: Bar Optimisation ─────────────────────────────────────────────
    ws5 = wb.create_sheet("Bar Optimisation")
    for col, w in zip("ABCDEFG", [30,8,14,14,14,14,12]):
        ws5.column_dimensions[col].width = w
    _xl_header(ws5, 1, 7, "BAR OPTIMISATION REPORT (FFD Heuristic)")
    row = 2
    for pk, bars in bar_data.items():
        if not bars:
            continue
        _xl_row_header(ws5, row, [f"Profile: {PROFILES[pk]['label']}",
                                   "Bar #", "Bar Len (mm)", "Used (mm)", "Waste (mm)", "Util %", "Cuts"], "C9922A")
        row += 1
        for b in bars:
            cuts_str = "; ".join(f"{c[0]:.0f}mm" for c in b.cuts[:6])
            if len(b.cuts) > 6:
                cuts_str += f" +{len(b.cuts)-6} more"
            shade = (b.bar_id % 2 == 0)
            _xl_data_row(ws5, row, [
                "", b.bar_id, b.bar_length, round(b.used, 1),
                round(b.waste, 1), f"{b.utilisation:.1f}%", cuts_str
            ], shade)
            row += 1
        row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
